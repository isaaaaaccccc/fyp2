from flask import Blueprint, render_template, request, flash, redirect, jsonify, get_flashed_messages, url_for, send_file, current_app
from werkzeug.utils import secure_filename
from application import db, bcrypt
from application.models import User
from flask_login import login_user, logout_user, login_required
from flask_wtf.csrf import CSRFProtect
from application.forms import CoachFilter, CoachDetails, DataUploadForm, BranchFilter, BranchForm
from application.models import DayOfWeek, User, Coach, Level, Branch, CoachBranch, CoachOffday, CoachPreference, Enrollment, PopularTimeslot, \
                            Timetable, TimetableEntry

from application.data_processor import load_database_driven
from application.enhanced_scheduler import EnhancedStrictConstraintScheduler, execute_enhanced_strict_constraint_scheduling
from application.util import transform_schedule_for_timetable_js, generate_sample_timetable

from collections import defaultdict
from datetime import datetime
import pandas as pd
from math import ceil

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter

api_bp = Blueprint('apis', __name__, url_prefix='/api')

@api_bp.route("/export-excel", methods=["POST"])
def export_excel():
    data = request.get_json() or {}
    if not data:
        return jsonify(error="No data provided"), 400

    CLASS_COLORS = {
        "L1":     "9FC5E8",
        "L2":     "FCE5CD",
        "L3":     "8E7CC3",
        "L4":     "F6B26B",
        "Flexi":  "D9EAD3",
        "Bubbly": "F4CCCC",
        "Jolly":  "00FFFF",
        "Tots":   "C2E7DA",
        "Lively": "FFF2CC"
    }

    TIME_SLOTS  = [f"{h:02d}{m:02d}" for h in range(9,21) for m in (0,30)]
    TIME_LABELS = [f"{h}:{m:02d}"     for h in range(9,21) for m in (0,30)]
    DAY_ORDER   = ["Monday","Tuesday","Wednesday","Thursday",
                   "Friday","Saturday","Sunday"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Timetable"

    # 1) Header row: merged day names
    col = 2
    for day in DAY_ORDER:
        span = len(data.get(day, {}))
        if span:
            ws.merge_cells(start_row=1, start_column=col,
                           end_row=1,   end_column=col+span-1)
            hdr = ws.cell(row=1, column=col, value=day)
            hdr.alignment = Alignment(horizontal="center")
            col += span

    # 2) Coach row
    ws.cell(row=2, column=1, value="Time")
    flat_pairs = [(d,c) for d in DAY_ORDER for c in data.get(d, {})]
    for idx, (_d, coach) in enumerate(flat_pairs):
        ws.cell(row=2, column=2+idx, value=coach)

    # 3) Time labels in col A
    for i, lbl in enumerate(TIME_LABELS, start=3):
        ws.cell(row=i, column=1, value=lbl)

    # 4) Fill sessions using duration as slot count
    for idx, (day, coach) in enumerate(flat_pairs):
        for s in data[day][coach]:
            t0 = str(s.get("start_time","")).replace(":", "").zfill(4)
            if t0 not in TIME_SLOTS:
                continue

            # treat duration directly as number of 30-min slots
            slot_count = int(s.get("duration", 0))
            if slot_count < 1:
                continue

            r1 = TIME_SLOTS.index(t0) + 3
            r2 = r1 + slot_count - 1
            c  = 2 + idx

            try:
                ws.merge_cells(start_row=r1, start_column=c,
                               end_row=r2,   end_column=c)
                cell = ws.cell(row=r1, column=c, value=s["name"])
                cell.fill = PatternFill(
                    "solid", fgColor="00"+CLASS_COLORS.get(s["name"], "FFFFFF")
                )
                cell.alignment = Alignment(
                    horizontal="center", vertical="center"
                )
            except ValueError:
                current_app.logger.warning(
                    f"Skipped overlapping session for {coach} on {day} at {t0}"
                )

    # 5) Layout tweaks
    ws.column_dimensions["A"].width = 12
    for i in range(2, 2 + len(flat_pairs)):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.freeze_panes = "A3"

    # 6) Return file
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(
        out,
        as_attachment=True,
        download_name="timetable.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ────────────────────────────────────────────────────────────
#  /api/export-coach-excel   (single coach, empty-safe)
# ────────────────────────────────────────────────────────────
@api_bp.route("/export-coach-excel", methods=["POST"])
def export_coach_excel():
    payload  = request.get_json() or {}
    schedule = payload.get("schedule", {})
    coaches  = payload.get("coaches", [])

    if len(coaches) != 1:
        return jsonify(error="Exactly one coach must be provided"), 400
    coach = coaches[0]

    CLASS_COLORS = {
        "L1":     "9FC5E8",
        "L2":     "FCE5CD",
        "L3":     "8E7CC3",
        "L4":     "F6B26B",
        "Flexi":  "D9EAD3",
        "Bubbly": "F4CCCC",
        "Jolly":  "00FFFF",
        "Tots":   "C2E7DA",
        "Lively": "FFF2CC"
    }

    TIME_SLOTS  = [f"{h:02d}{m:02d}" for h in range(9,21) for m in (0,30)]
    TIME_LABELS = [f"{h}:{m:02d}"     for h in range(9,21) for m in (0,30)]
    DAY_ORDER   = ["Monday","Tuesday","Wednesday","Thursday",
                   "Friday","Saturday","Sunday"]

    flat_days = [d for d in DAY_ORDER if coach in schedule.get(d, {})]

    wb = Workbook()
    ws = wb.active
    ws.title = coach

    if not flat_days:
        ws["A1"] = "No sessions"
    else:
        # header row for days
        col = 2
        for day in flat_days:
            ws.merge_cells(start_row=1, start_column=col,
                           end_row=1,   end_column=col)
            hdr = ws.cell(row=1, column=col, value=day)
            hdr.alignment = Alignment(horizontal="center")
            col += 1

        # coach row
        ws.cell(row=2, column=1, value="Time")
        for idx in range(len(flat_days)):
            ws.cell(row=2, column=2+idx, value=coach)

        # time labels
        for i, lbl in enumerate(TIME_LABELS, start=3):
            ws.cell(row=i, column=1, value=lbl)

        # fill sessions
        for idx, day in enumerate(flat_days):
            for s in schedule[day][coach]:
                t0 = str(s.get("start_time","")).replace(":", "").zfill(4)
                if t0 not in TIME_SLOTS:
                    continue

                slot_count = int(s.get("duration", 0))
                if slot_count < 1:
                    continue

                r1 = TIME_SLOTS.index(t0) + 3
                r2 = r1 + slot_count - 1
                c  = 2 + idx

                try:
                    ws.merge_cells(start_row=r1, start_column=c,
                                   end_row=r2,   end_column=c)
                    cell = ws.cell(row=r1, column=c, value=s["name"])
                    cell.fill = PatternFill(
                        "solid", fgColor="00"+CLASS_COLORS.get(s["name"], "FFFFFF")
                    )
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                except ValueError:
                    current_app.logger.warning(
                        f"Skipped overlapping for {coach} on {day} at {t0}"
                    )

    # layout tweaks
    ws.column_dimensions["A"].width = 12
    for i in range(2, 2 + max(1, len(flat_days))):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.freeze_panes = "A3"

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(
        out,
        as_attachment=True,
        download_name=f"{coach}_timetable.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@api_bp.route('/coach', methods=['GET'])
def get_coach():
    form = CoachFilter(request.args)
    query = db.session.query(Coach)

    if form.name.data:
        query = query.filter(Coach.name.ilike(f"%{form.name.data}%"))
    if form.branch.data:
        query = query.join(Coach.assigned_branches).filter(CoachBranch.branch_id == form.branch.data.id)
    if form.position.data:
        query = query.filter(Coach.position == form.position.data)
    if form.level.data:
        query = query.join(Coach.preferred_levels).filter_by(level_id=form.level.data.id)

    coaches = query.all()

    return jsonify([{
        'id': coach.id,
        'name': coach.name,
        'residential_area': coach.residential_area,
        'position': coach.position,
        'status': coach.status,
        'assigned_branches': [{
            'id': cb.branch.id,
            'name': cb.branch.name,
            'abbrv': cb.branch.abbrv,
        } for cb in coach.assigned_branches],
        'offdays': [{
            'day': DayOfWeek(cd.day).name,
            'am': cd.am,
            'reason': cd.reason
        } for cd in coach.offdays],
        'preferred_levels': [{
            'id': cl.level.id,
            'name': cl.level.name,
            'alias': cl.level.alias,
        } for cl in coach.preferred_levels]
    } for coach in coaches])

@api_bp.route('/coach/<int:id>', methods=['GET'])
def get_coach_by_id(id):
    coach = db.session.query(Coach).filter(Coach.id == id).first()

    return jsonify({
        'id': coach.id,
        'name': coach.name,
        'residential_area': coach.residential_area,
        'position': coach.position,
        'status': coach.status,
        'assigned_branches': [{
            'id': cb.branch.id,
            'name': cb.branch.name,
            'abbrv': cb.branch.abbrv,
        } for cb in coach.assigned_branches],
        'offdays': [{
            'day': DayOfWeek(cd.day).name,
            'am': cd.am,
            'reason': cd.reason
        } for cd in coach.offdays],
        'preferred_levels': [{
            'id': cl.level.id,
            'name': cl.level.name,
            'alias': cl.level.alias,
        } for cl in coach.preferred_levels]
    })

@api_bp.route('/coach/<int:id>', methods=['PUT'])
def update_coach_by_id(id):
    data = request.get_json()
    form = CoachDetails(data=data)

    if not form.validate():
        return jsonify({'error': form.errors}), 400
    
    coach = Coach.query.get_or_404(id)

    coach.name = form.editName.data
    coach.residential_area = form.editResidence.data
    coach.position = form.editPosition.data
    coach.status = 'Part time' if form.editPosition.data == 'Part time' else 'Full time'
    
    CoachBranch.query.filter_by(coach_id=id).delete()
    for branchField in form.editBranch:
        if not branchField.data:
            continue

        cb = CoachBranch(coach_id=id, branch_id=branchField.id.removeprefix('editBranch_'))
        db.session.add(cb)

    CoachPreference.query.filter_by(coach_id=id).delete()
    for levelField in form.editLevel:
        if not levelField.data:
            continue
            
        cl = CoachPreference(coach_id=id, level_id=levelField.id.removeprefix('editLevel_'))
        db.session.add(cl)
    
    CoachOffday.query.filter_by(coach_id=id).delete()
    for row in form.editOffday:
        for offdayField in row:
            if not offdayField.data:
                continue

            day, am = offdayField.id.removeprefix('editOffday_').split('_')
            cd = CoachOffday(coach_id=id, day=DayOfWeek[day.upper()].value, am=int(am == 'AM'))
            db.session.add(cd)

    db.session.commit()
    
    return get_coach_by_id(id), 200

@api_bp.route('/coach/<int:id>', methods=['DELETE'])
def delete_coach_by_id(id):
    coach = Coach.query.get_or_404(id)

    db.session.delete(coach)
    db.session.commit()
    
    return '', 204

@api_bp.route('/branch/', methods=['GET'])
def get_branch():
    """Get all branches"""
    try:
        branches = Branch.query.all()
        branches_data = []
        
        for branch in branches:
            created_at = datetime.now()
            if hasattr(branch, 'created_at') and branch.created_at:
                created_at = branch.created_at
                
            branches_data.append({
                'id': branch.id,
                'name': branch.name,
                'abbrv': branch.abbrv,
                'max_classes': branch.max_classes,
                'created_at': created_at.isoformat()
            })
        
        print(f"Returning {len(branches_data)} branches")
        return jsonify({
            'success': True,
            'branches': branches_data
        })
    except Exception as e:
        print(f"Error getting branches: {e}")
        return jsonify({
            'success': False,
            'message': f"Error getting branches: {str(e)}"
        }), 500

@api_bp.route('/branch/<int:branch_id>', methods=['GET'])
def get_branch_by_id(branch_id):
    """Get a specific branch"""
    try:
        branch = Branch.query.get(branch_id)
        if not branch:
            return jsonify({
                'success': False,
                'message': f"Branch with ID {branch_id} not found"
            }), 404
            
        branch_data = {
            'id': branch.id,
            'name': branch.name,
            'abbrv': branch.abbrv,
            'max_classes': branch.max_classes,
            'created_at': datetime.now().isoformat() if not hasattr(branch, 'created_at') else branch.created_at.isoformat()
        }
        
        return jsonify({
            'success': True,
            'branch': branch_data
        })
    except Exception as e:
        print(f"Error getting branch: {e}")
        return jsonify({
            'success': False,
            'message': f"Error getting branch: {str(e)}"
        }), 500

@api_bp.route('/branch/', methods=['POST'])
def create_branch():
    """Create a new branch"""
    try:
        data = request.get_json()
        
        # Validate required fields
        if not data or 'name' not in data or 'abbrv' not in data:
            return jsonify({
                'success': False,
                'message': "Missing required fields: name, abbrv"
            }), 400
            
        # Check if branch with same name or abbreviation already exists
        if Branch.query.filter_by(name=data['name']).first():
            return jsonify({
                'success': False,
                'message': f"Branch with name '{data['name']}' already exists"
            }), 400
            
        if Branch.query.filter_by(abbrv=data['abbrv'].upper()).first():
            return jsonify({
                'success': False,
                'message': f"Branch with abbreviation '{data['abbrv'].upper()}' already exists"
            }), 400
        
        # Create new branch
        new_branch = Branch(
            name=data['name'],
            abbrv=data['abbrv'].upper(),
            max_classes=data.get('max_classes', 4)
        )
        
        db.session.add(new_branch)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f"Branch '{data['name']}' created successfully",
            'branch': {
                'id': new_branch.id,
                'name': new_branch.name,
                'abbrv': new_branch.abbrv,
                'max_classes': new_branch.max_classes,
                'created_at': datetime.now().isoformat()
            }
        })
    except Exception as e:
        db.session.rollback()
        print(f"Error creating branch: {e}")
        return jsonify({
            'success': False,
            'message': f"Error creating branch: {str(e)}"
        }), 500

@api_bp.route('/branch/<int:branch_id>', methods=['PUT'])
def update_branch(branch_id):
    """Update an existing branch"""
    try:
        # Try to parse JSON data - add explicit error handling
        try:
            data = request.get_json()
            if data is None:
                print(f"No JSON data in request or invalid JSON format")
                return jsonify({
                    'success': False,
                    'message': "No JSON data provided or invalid JSON format"
                }), 400
        except Exception as json_error:
            print(f"JSON parsing error: {str(json_error)}")
            return jsonify({
                'success': False,
                'message': f"Invalid JSON data: {str(json_error)}"
            }), 400
            
        # Log received data
        print(f"Received update request for branch {branch_id}: {data}")
        
        branch = Branch.query.get(branch_id)
        if not branch:
            print(f"Branch with ID {branch_id} not found")
            return jsonify({
                'success': False,
                'message': f"Branch with ID {branch_id} not found"
            }), 404
        
        # Check required fields
        missing_fields = []
        if 'name' not in data or not data['name']:
            missing_fields.append('name')
        if 'abbrv' not in data or not data['abbrv']:
            missing_fields.append('abbrv')
        if 'max_classes' not in data:
            missing_fields.append('max_classes')
            
        if missing_fields:
            print(f"Missing required fields: {', '.join(missing_fields)}")
            return jsonify({
                'success': False,
                'message': f"Missing required fields: {', '.join(missing_fields)}"
            }), 400
        
        # Update branch with simple approach
        try:
            branch.name = str(data['name']).strip()
            branch.abbrv = str(data['abbrv']).strip().upper()
            branch.max_classes = int(data['max_classes'])
            
            # Print branch data before commit
            print(f"Updated branch data - ID: {branch.id}, Name: {branch.name}, Abbrv: {branch.abbrv}, Max Classes: {branch.max_classes}")
            
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': f"Branch '{branch.name}' updated successfully",
                'branch': {
                    'id': branch.id,
                    'name': branch.name,
                    'abbrv': branch.abbrv,
                    'max_classes': branch.max_classes,
                    'created_at': datetime.now().isoformat()
                }
            })
        except Exception as update_error:
            db.session.rollback()
            print(f"Error updating branch: {str(update_error)}")
            return jsonify({
                'success': False,
                'message': f"Error updating branch: {str(update_error)}"
            }), 500
            
    except Exception as e:
        print(f"Unexpected error in update_branch: {str(e)}")
        return jsonify({
            'success': False,
            'message': f"Server error: {str(e)}"
        }), 500

@api_bp.route('/branch/<int:branch_id>', methods=['DELETE'])
def delete_branch(branch_id):
    """Delete a branch"""
    try:
        branch = Branch.query.get(branch_id)
        if not branch:
            return jsonify({
                'success': False,
                'message': f"Branch with ID {branch_id} not found"
            }), 404
        
        # Check if branch is associated with any coaches
        if CoachBranch.query.filter_by(branch_id=branch_id).count() > 0:
            return jsonify({
                'success': False,
                'message': f"Cannot delete branch '{branch.name}' because it is associated with coaches"
            }), 400
        
        branch_name = branch.name
        db.session.delete(branch)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f"Branch '{branch_name}' deleted successfully"
        })
    except Exception as e:
        db.session.rollback()
        print(f"Error deleting branch: {e}")
        return jsonify({
            'success': False,
            'message': f"Error deleting branch: {str(e)}"
        }), 500


#############
# TIMETABLE #
#############

@api_bp.route('/timetable/generate/', methods=['POST'])
def generate():
    """Generate timetable data for frontend visualization using the database"""
    try:
        print("Starting timetable generation...")

        config = request.get_json() or {}
        print(config)
        
        # Step 1: Load data from database using data_processor
        print("Loading data from database...")
        data = load_database_driven()
        
        if not data:
            print("Failed to load data from database")
            return jsonify({
                'success': False,
                'message': 'Failed to load data from database',
                'error': 'Database data unavailable or incomplete'
            }), 500
            
        # Log the loaded data stats
        coaches_count = len(data.get('coaches_data', {}))
        enrollment_count = len(data.get('enrollment_dict', {}))
        assignment_count = len(data.get('feasible_assignments', []))
        
        print(f"Data loaded: {coaches_count} coaches, {enrollment_count} enrollments, {assignment_count} feasible assignments")
        
        if coaches_count == 0 or enrollment_count == 0 or assignment_count == 0:
            print("Insufficient data for scheduling")
            return jsonify({
                'success': False,
                'message': 'Insufficient data for scheduling'
            }), 400
        
        # Step 2: Run the scheduling algorithm from enhanced_scheduler
        print("Running enhanced strict constraint scheduling...")
        scheduler = EnhancedStrictConstraintScheduler(data, config)
        results = scheduler.schedule_with_complete_coverage()
        # results = execute_enhanced_strict_constraint_scheduling(data)
        
        if not results or 'schedule' not in results or not results['schedule']:
            print("Scheduler failed to generate a timetable")
            return jsonify({
                'success': False,
                'message': 'Scheduler failed to generate a timetable'
            }), 500 
        
        
        schedule_count = len(results['schedule'])
        coverage = results['statistics'].get('coverage_percentage', 0)
        print(f"Schedule generated with {schedule_count} classes ({coverage:.1f}% coverage)")
        
        # Step 3: Convert the schedule to the format expected by timetable.js
        processed_data = transform_schedule_for_timetable_js(results['schedule'])
        
        return jsonify(processed_data)
        
    except Exception as e:
        import traceback
        error_msg = f"Error generating timetable: {str(e)}"
        print(f"{error_msg}")
        traceback.print_exc()
        
        print("Exception occurred - falling back to sample timetable")
        return jsonify({
            'success': False,
            'message': 'Something went wrong.'
        }), 500

@api_bp.route('/timetable/save/', methods=['POST'])
def save_timetable():
    timetable_data = request.get_json()

    if not timetable_data:
        return jsonify({'success': False, 'message': 'No data provided'}), 400
    
    timetable = Timetable()
    db.session.add(timetable)
    db.session.flush()  # Ensure it is generated before it is used

    for branch, branch_data in timetable_data.items():
        branch_id = Branch.query.filter(Branch.abbrv == branch).first().id

        for day, schedule in branch_data['schedule'].items():
            day_int = DayOfWeek[day[:3].upper()].value

            for coach, classes in schedule.items():
                coach_id = Coach.query.filter(Coach.name == coach).first().id

                for details in classes:
                    level = details['name']
                    start_time = datetime.strptime(details['start_time'], "%H%M").time()
                    level_id = Level.query.filter(Level.alias == level).first().id

                    entry = TimetableEntry(
                        timetable_id = timetable.id,
                        branch_id=branch_id,
                        coach_id=coach_id,
                        level_id=level_id,
                        start_time=start_time,
                        day=day_int
                    )
                    db.session.add(entry)
        
    db.session.commit()
    return jsonify({
        'success': True,
        'message': 'Timetable saved',
        'timetable_id': timetable.id
    }), 201


def format_timetable(timetable):
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    data = {}

    for entry in timetable.entries:
        branch, coach, level = entry.branch, entry.coach, entry.level
        start_time = entry.start_time.strftime("%H%M")
        day = days[entry.day]

        if branch.abbrv not in data:
            data[branch.abbrv] = {
                "coaches": [],
                "schedule": {}
            }

        if coach.name not in data[branch.abbrv]["coaches"]:
            data[branch.abbrv]["coaches"].append(coach.name)

        if day not in data[branch.abbrv]["schedule"]:
            data[branch.abbrv]["schedule"][day] = {}

        if coach.name not in data[branch.abbrv]["schedule"][day]:
            data[branch.abbrv]["schedule"][day][coach.name] = []

        data[branch.abbrv]["schedule"][day][coach.name].append({
            "duration": level.duration,
            "name": level.alias,
            "start_time": start_time
        })

    return {
        'id': timetable.id,
        'date_created': timetable.date_created.isoformat(),
        'active': bool(timetable.active),
        'data': data
    }


@api_bp.route('/timetable/', methods=['GET'])
def get_timetable():
    # results_per_page = int(request.args.get('results', 5))
    results_per_page = int(request.args.get('results', 100))  # TODO: Change back to 5 once pagination system is ready
    page = int(request.args.get('page', 1))
    show_active = bool(request.args.get('show_active', False))

    total_count = Timetable.query.count()
    max_pages = ceil(total_count / results_per_page)

    if total_count == 0:
        return jsonify({
            "results": [],
            "page": page,
            "max_pages": max_pages,
            "total_count": total_count
        }), 200

    if not 1 <= results_per_page <= 100:
        return jsonify({
            'success': False,
            'message': f'Invalid results per page. Must be between 1 and 20.'
        }), 400

    if not 1 <= page <= max_pages:
        return jsonify({
            'success': False,
            'message': f'Invalid page. Must be between 1 and {max_pages}'
        }), 400

    timetables = Timetable.query \
        .order_by(Timetable.date_created.desc()) \
        .offset((page - 1) * results_per_page) \
        .limit(results_per_page) \
        .all()
    
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    response = [format_timetable(t) for t in timetables]

    if show_active:
        active = Timetable.query.filter(Timetable.active==True).first()
        if active and all(t['id'] != active.id for t in response):
            response.insert(0, format_timetable(active))

    return jsonify({
        "results": response,
        "page": page,
        "max_pages": max_pages,
        "total_count": total_count
    }), 200


@api_bp.route('/timetable/<int:id>', methods=['GET'])
def get_timetable_by_id(id):
    timetable = Timetable.query.get_or_404(id)
    
    return jsonify(format_timetable(timetable)), 200

@api_bp.route('/timetable/<int:id>', methods=['DELETE'])
def delete_timetable(id):
    timetable = Timetable.query.get_or_404(id)
    db.session.delete(timetable)

    db.session.commit()

    return jsonify({
        'success': True,
        'message': f'Deleted timetable id {id}.'
    }), 200

@api_bp.route('/timetable/<int:id>/activate', methods=['POST'])
def set_active_timetable(id):
    timetable = Timetable.query.get_or_404(id)
    Timetable.query.update({Timetable.active: False})
    db.session.flush()

    timetable.active = True
    db.session.commit()

    return jsonify({
        'success': True,
        'message': 'Updated active timetable.'
    }), 200

@api_bp.route('/timetable/active', methods=['GET'])
def get_active_timetable():
    timetable = Timetable.query.filter(Timetable.active == True).first()

    if not timetable:
        return jsonify({
            'success': False,
            'message': 'No active timetable found.'
        }), 404

    return get_timetable_by_id(timetable.id)

@api_bp.route('/coach/', methods=['GET'])
def get_all_coaches():
    """Get all coaches in a format suitable for the timetable interface"""
    try:
        # Query all coaches from the database
        coaches = Coach.query.all()
        
        # Format the response
        coach_list = []
        for coach in coaches:
            coach_data = {
                'name': coach.name
            }
            
            # Add status if it exists in your model (your existing code uses this field)
            if hasattr(coach, 'status'):
                coach_data['status'] = coach.status
                
            coach_list.append(coach_data)
        
        return jsonify({
            'success': True,
            'coaches': coach_list
        })
    except Exception as e:
        print(f"Error fetching coaches: {str(e)}")
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500